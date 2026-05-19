from pathlib import Path
import re
import sys

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from server.company_db import get_product, list_products, set_product_details


WORKBOOK_PATH = Path("/home/cybertechna/Downloads/perfume_inventory_verified.xlsx")
BATCH_NOTE = "Imported from supplier line items on 2026-04-02"


def normalize_name(value):
    text = (value or "").strip().lower()
    text = re.sub(r"\([^)]*\)", "", text)
    text = text.replace("&", "and")
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def parse_money(value):
    if value is None:
        return None
    text = str(value).strip().replace("$", "").replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def main():
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["Perfume Inventory"]

    inventory_rows = [
        row for row in list_products(active_only=False, low_stock_only=False)
        if (row.get("supplier_name") == "Supplier Batch Import" and row.get("notes") == BATCH_NOTE)
    ]
    by_norm = {normalize_name(row.get("name")): row for row in inventory_rows}

    updated = 0
    unmatched_sheet_rows = []
    matched_ids = []

    for r in range(2, ws.max_row + 1):
        product_name = ws.cell(r, 2).value
        if not product_name:
            continue
        normalized = normalize_name(product_name)
        target = by_norm.get(normalized)
        if not target:
            unmatched_sheet_rows.append(product_name)
            continue

        barcode = ws.cell(r, 3).value
        retail_price = parse_money(ws.cell(r, 4).value)
        top_notes = ws.cell(r, 5).value
        mid_notes = ws.cell(r, 6).value
        base_notes = ws.cell(r, 7).value
        dupe = ws.cell(r, 8).value
        image_url = ws.cell(r, 9).value
        verification = ws.cell(r, 10).value

        existing = get_product(int(target["id"])) or {}
        note_bits = [BATCH_NOTE]
        if verification:
            note_bits.append(str(verification).strip())
        notes_value = " | ".join(note_bits)

        set_product_details(
            int(target["id"]),
            name=product_name,
            barcode=str(barcode).strip() if barcode else existing.get("barcode"),
            retail_price=retail_price if retail_price is not None else existing.get("retail_price"),
            note_top=top_notes,
            note_mid=mid_notes,
            note_base=base_notes,
            dupe_inspiration=dupe,
            dupe_notes=str(verification).strip() if verification else existing.get("dupe_notes"),
            media_url=image_url,
            notes=notes_value,
        )
        matched_ids.append(int(target["id"]))
        updated += 1

    print(
        {
            "ok": True,
            "updated_products": updated,
            "matched_ids_preview": matched_ids[:10],
            "unmatched_sheet_rows": unmatched_sheet_rows[:20],
        }
    )


if __name__ == "__main__":
    main()
