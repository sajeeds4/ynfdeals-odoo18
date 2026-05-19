#!/usr/bin/env python3
"""
Enrich TikTok LIVE auction orders export with local inventory product info.

Input assumptions (based on your sheet):
- 'lot no' is the seller SKU/lot number (we keep it as-is)
- 'barcode' is the barcode for the sold item
- 'price' is the actual sold price (we DO NOT modify it)

We append product details from the Postgres `products` table matched by barcode.

Usage:
  ./.venv-excel/bin/python tools/enrich_tiktok_orders.py \
    --in  /home/cybertechna/Downloads/tiktokday2_with_barcodes.xlsx \
    --out /home/cybertechna/Downloads/tiktokday2_enriched.xlsx \
    --cancel-lots "31,32,26"
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from server.config import POSTGRES_SIDECAR_SCHEMA
from server.postgres_cutover import _pg_connect, ensure_wave1_postgres_schema, postgres_available


def norm_barcode(v: Any) -> str:
  if v is None:
    return ""
  if isinstance(v, str):
    return v.strip()
  if isinstance(v, (int,)):
    return str(v)
  if isinstance(v, float):
    # Excel sometimes stores big integers as float; avoid scientific notation.
    return f"{v:.0f}"
  return str(v).strip()


def load_products_by_barcode() -> Dict[str, dict]:
  if not postgres_available():
    raise SystemExit("postgres_required: enrich_tiktok_orders no longer reads SQLite whatnot.db")
  ensure_wave1_postgres_schema()
  con = _pg_connect()
  cur = con.cursor()
  cur.execute(
    f"""
    SELECT
      barcode,
      id,
      name,
      sku,
      brand,
      cost_price,
      retail_price,
      on_hand_qty,
      active,
      gender,
      supplier_name,
      storage_bin
    FROM {POSTGRES_SIDECAR_SCHEMA}.products
    WHERE barcode IS NOT NULL AND barcode != ''
    """
  )
  columns = [desc[0] for desc in cur.description]
  out: Dict[str, dict] = {}
  for row in cur.fetchall():
    product = dict(zip(columns, row))
    bc = (product["barcode"] or "").strip()
    if not bc:
      continue
    out[bc] = product
  con.close()
  return out


def ensure_columns(ws, headers, new_cols):
  for c in new_cols:
    if c not in headers:
      headers.append(c)
      ws.cell(row=1, column=len(headers), value=c)
  return headers


def main() -> None:
  ap = argparse.ArgumentParser()
  ap.add_argument("--in", dest="inp", required=True, help="Input xlsx path")
  ap.add_argument("--out", dest="out", required=True, help="Output xlsx path")
  ap.add_argument("--db", dest="db", default="", help="Legacy ignored SQLite path; products are read from Postgres")
  ap.add_argument("--cancel-lots", dest="cancel_lots", default="", help="Comma-separated lot numbers to mark cancelled")
  ap.add_argument("--out-non-cancelled", dest="out_non_cancelled", default="", help="Optional output xlsx excluding cancelled lots")
  args = ap.parse_args()

  if not os.path.exists(args.inp):
    raise SystemExit(f"Input not found: {args.inp}")
  products = load_products_by_barcode()

  wb = load_workbook(args.inp)
  ws = wb.active

  headers = [c.value for c in ws[1]]
  headers = [h.strip() if isinstance(h, str) else (h or "") for h in headers]
  idx = {h: i + 1 for i, h in enumerate(headers) if h}

  if "barcode" not in idx:
    raise SystemExit("Input sheet missing required column: barcode")

  cancel_set = set()
  if args.cancel_lots.strip():
    for part in args.cancel_lots.split(","):
      part = part.strip()
      if not part:
        continue
      try:
        cancel_set.add(int(part))
      except ValueError:
        # keep non-int lots as raw strings too
        cancel_set.add(part)

  # Append new columns.
  new_cols = [
    "product_id",
    "product_name",
    "inventory_sku",
    "brand",
    "cost_price",
    "retail_price",
    "on_hand_qty",
    "active",
    "gender",
    "supplier_name",
    "storage_bin",
    "match_status",
    "is_cancelled",
  ]
  headers = ensure_columns(ws, headers, new_cols)
  idx = {h: i + 1 for i, h in enumerate(headers) if h}

  for r in range(2, ws.max_row + 1):
    bc = norm_barcode(ws.cell(row=r, column=idx["barcode"]).value)
    lot_val = ws.cell(row=r, column=idx.get("lot no", 0)).value if "lot no" in idx else None
    is_cancelled = False
    if cancel_set:
      if isinstance(lot_val, (int, float)) and int(lot_val) in cancel_set:
        is_cancelled = True
      elif lot_val in cancel_set:
        is_cancelled = True
    ws.cell(row=r, column=idx["is_cancelled"], value=1 if is_cancelled else 0)

    if not bc:
      ws.cell(row=r, column=idx["match_status"], value="missing_barcode")
      continue

    p: Optional[dict] = products.get(bc)
    if not p:
      ws.cell(row=r, column=idx["match_status"], value="barcode_not_found")
      continue

    ws.cell(row=r, column=idx["product_id"], value=p.get("id"))
    ws.cell(row=r, column=idx["product_name"], value=p.get("name"))
    ws.cell(row=r, column=idx["inventory_sku"], value=p.get("sku"))
    ws.cell(row=r, column=idx["brand"], value=p.get("brand"))
    ws.cell(row=r, column=idx["cost_price"], value=p.get("cost_price"))
    ws.cell(row=r, column=idx["retail_price"], value=p.get("retail_price"))
    ws.cell(row=r, column=idx["on_hand_qty"], value=p.get("on_hand_qty"))
    ws.cell(row=r, column=idx["active"], value=p.get("active"))
    ws.cell(row=r, column=idx["gender"], value=p.get("gender"))
    ws.cell(row=r, column=idx["supplier_name"], value=p.get("supplier_name"))
    ws.cell(row=r, column=idx["storage_bin"], value=p.get("storage_bin"))
    ws.cell(row=r, column=idx["match_status"], value="ok")

  os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)

  # If cancelled lots are provided but they don't exist in the export, keep an audit sheet.
  if cancel_set:
    # Ensure we don't duplicate if rerun on an already-enriched workbook.
    existing = set(wb.sheetnames)
    title = "Cancelled Lots"
    if title in existing:
      del wb[title]
    ws_c = wb.create_sheet(title)
    ws_c.append(["lot no", "note"])
    for lot in sorted(cancel_set, key=lambda x: (isinstance(x, str), x)):
      ws_c.append([lot, "Cancelled (not present in export)" ])

  wb.save(args.out)

  print(f"Wrote: {args.out}")

  if args.out_non_cancelled:
    # Create a copy workbook excluding cancelled rows.
    wb2 = load_workbook(args.out)
    ws2 = wb2.active
    headers2 = [c.value for c in ws2[1]]
    idx2 = {h: i + 1 for i, h in enumerate(headers2) if h}
    if "is_cancelled" not in idx2:
      raise SystemExit("Internal error: is_cancelled column missing")

    # Delete rows bottom-up to keep indices stable.
    removed = 0
    for rr in range(ws2.max_row, 1, -1):
      v = ws2.cell(row=rr, column=idx2["is_cancelled"]).value
      if v in (1, "1", True):
        ws2.delete_rows(rr, 1)
        removed += 1

    os.makedirs(os.path.dirname(args.out_non_cancelled) or ".", exist_ok=True)
    wb2.save(args.out_non_cancelled)
    print(f"Wrote: {args.out_non_cancelled} (removed {removed} cancelled rows)")


if __name__ == "__main__":
  main()
